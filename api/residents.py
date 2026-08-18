"""住客主档与 Excel 批次汇入。

汇入的设计要点：校验失败的列不写库，收集成错误清单一并回传，
让使用者知道第几列为什么没进去，而不是整批失败或静默丢弃。
"""

import io
from datetime import datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from core import audit, db, deps
from domain import rules, seed

router = APIRouter(tags=["residents"])

COLUMNS = ["床號", "姓名", "性別", "資助類別", "服務軌道", "每週目標節數", "上次評估日期"]
VALID_FUNDING = set(seed.FUNDING_TYPES)
VALID_TRACKS = set(seed.SERVICE_TRACKS)
VALID_GENDER = {"男", "女"}


@router.get("/residents")
def list_residents(user: deps.User = Depends(deps.current_user)) -> dict:
    where, args = deps.facility_scope(user, "r")
    rows = db.query(
        "SELECT r.*, f.name AS facility_name FROM residents r"
        " JOIN facilities f ON f.id = r.facility_id"
        " WHERE r.is_deleted = 0" + where +
        " ORDER BY r.facility_id, r.bed_no",
        args,
    )
    for row in rows:
        row["assessment"] = rules.assessment_status(row["last_assessed"])
    return {"scope_label": user.scope_label, "items": rows}


@router.get("/residents/template")
def download_template() -> StreamingResponse:
    """产生汇入模板。表头即欄位映射依据。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "住客名單"
    ws.append(COLUMNS)
    ws.append(["A101", "陳大文", "男", "甲一買位", "資助復康", 3, "2026-05-01"])
    for index, _ in enumerate(COLUMNS, start=1):
        ws.column_dimensions[chr(64 + index)].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="residents_template.xlsx"'},
    )


def _validate(row: dict, line: int, seen_beds: set) -> str | None:
    if not row["bed_no"]:
        return "床號為空"
    if not row["name"]:
        return "姓名為空"
    if row["bed_no"] in seen_beds:
        return f"床號 {row['bed_no']} 在檔案內重複"
    if row["gender"] not in VALID_GENDER:
        return f"性別「{row['gender']}」不是 男 或 女"
    if row["funding_type"] not in VALID_FUNDING:
        return f"資助類別「{row['funding_type']}」不在允許範圍"
    if row["service_track"] not in VALID_TRACKS:
        return f"服務軌道「{row['service_track']}」不在允許範圍"
    if row["weekly_target"] is None:
        return "每週目標節數不是數字"
    if row["last_assessed"]:
        try:
            rules.parse_date(row["last_assessed"])
        except ValueError:
            return f"上次評估日期「{row['last_assessed']}」格式不正確，應為 YYYY-MM-DD"
    return None


def _cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


@router.post("/residents/import")
async def import_residents(
    file: UploadFile = File(...),
    user: deps.User = Depends(deps.require("supervisor", "admin")),
) -> dict:
    from openpyxl import load_workbook

    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "請上傳 .xlsx 檔案")

    try:
        wb = load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception:
        raise HTTPException(400, "檔案無法解析，請確認是有效的 Excel 檔")

    ws = wb.active
    header = [_cell(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index_of = {name: i for i, name in enumerate(header)}
    missing = [c for c in COLUMNS[:5] if c not in index_of]
    if missing:
        raise HTTPException(400, "缺少必要欄位：" + "、".join(missing))

    existing = {
        r["bed_no"] for r in db.query(
            "SELECT bed_no FROM residents WHERE facility_id = ? AND is_deleted = 0",
            (user.facility_id,))
    }

    accepted: list[tuple] = []
    errors: list[dict] = []
    seen_beds: set[str] = set()
    now = datetime.now().isoformat(timespec="seconds")

    for line, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if raw is None or all(v is None for v in raw):
            continue

        def get(column: str) -> str:
            position = index_of.get(column)
            return _cell(raw[position]) if position is not None and position < len(raw) else ""

        try:
            target = int(float(get("每週目標節數") or 3))
        except ValueError:
            target = None

        row = {
            "bed_no": get("床號"),
            "name": get("姓名"),
            "gender": get("性別"),
            "funding_type": get("資助類別"),
            "service_track": get("服務軌道"),
            "weekly_target": target,
            "last_assessed": get("上次評估日期") or None,
        }

        problem = _validate(row, line, seen_beds)
        if problem is None and row["bed_no"] in existing:
            problem = f"床號 {row['bed_no']} 在系統中已存在"

        if problem:
            errors.append({"line": line, "bed_no": row["bed_no"], "reason": problem})
            continue

        seen_beds.add(row["bed_no"])
        accepted.append((
            row["bed_no"], row["name"], row["gender"], row["funding_type"], 0,
            row["weekly_target"], row["service_track"], row["last_assessed"],
            user.facility_id, 0, now,
        ))

    if accepted:
        db.execute_many(
            "INSERT INTO residents(bed_no, name, gender, funding_type, sc_flag,"
            " weekly_target, service_track, last_assessed, facility_id, is_deleted, created_at)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            accepted,
        )

    reasons = "；".join(f"第 {e['line']} 列 {e['reason']}" for e in errors[:3])
    if len(errors) > 3:
        reasons += f"；另有 {len(errors) - 3} 列"
    audit.write(
        user, "匯入住客名單",
        f"匯入住客名單「{file.filename}」，成功 {len(accepted)} 筆、失敗 {len(errors)} 筆",
        "resident", file.filename,
        details=[
            audit.change("檔案名稱", after=file.filename),
            audit.change("成功筆數", after=len(accepted)),
            audit.change("失敗筆數", after=len(errors)),
            audit.change("失敗原因", after=reasons or "無"),
        ],
    )

    return {
        "filename": file.filename,
        "imported": len(accepted),
        "failed": len(errors),
        "errors": errors[:50],
    }


@router.delete("/residents/{resident_id}")
def delete_resident(
    resident_id: int,
    user: deps.User = Depends(deps.require("supervisor", "admin")),
) -> dict:
    """软删除。资料不真的消失，只是标记，历史节次与表单的关联仍然完整。

    受监管的场景里硬删除是不能接受的：稽查时被问到「这位住客的纪录呢」，
    答不出来比答错更糟。所以系统全篇没有真正的 DELETE。
    """
    where, args = deps.facility_scope(user)
    row = db.query_one(
        "SELECT * FROM residents WHERE id = ? AND is_deleted = 0" + where,
        [resident_id] + args,
    )
    if not row:
        raise HTTPException(404, "找不到此住客，或不在你的資料範圍內")

    db.execute("UPDATE residents SET is_deleted = 1 WHERE id = ?", (resident_id,))
    audit.write(
        user, "移除住客",
        f"將 {row['name']}（{row['bed_no']}）從名單移除",
        "resident", resident_id,
        details=[
            audit.change("狀態", before="在住", after="已移除"),
            audit.change("床號", after=row["bed_no"]),
            audit.change("資助類別", after=row["funding_type"]),
            audit.change("移除方式", after="軟刪除，歷史紀錄保留可查"),
        ],
    )
    return {"id": resident_id, "name": row["name"], "bed_no": row["bed_no"]}
